"""
Excel Document Processor
Xử lý file Excel (.xlsx, .xls), trích xuất dữ liệu từ các sheet
"""

import logging
from typing import Dict, List, Any, Optional
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import xlrd
import sys
from pathlib import Path

# Add project directories to path
current_dir = Path(__file__).parent
load_documents_dir = current_dir.parent
sys.path.append(str(load_documents_dir))

from .base_processor import BaseDocumentProcessor, DocumentContent, PageContent
from controllers.rag.load_documents.utils import DocumentUtils

logger = logging.getLogger(__name__)

class ExcelProcessor(BaseDocumentProcessor):
    """
    Processor cho file Excel
    Sử dụng pandas, openpyxl để trích xuất dữ liệu
    """
    
    def __init__(self):
        super().__init__()
        self.supported_extensions = ['.xlsx', '.xls']
    
    def can_process(self, file_extension: str) -> bool:
        """Kiểm tra có thể xử lý file Excel không"""
        return file_extension.lower() in self.supported_extensions
    
    async def extract_content(self, file_path: str, file_data: bytes) -> DocumentContent:
        """
        Trích xuất nội dung từ file Excel
        
        Args:
            file_path: Đường dẫn file
            file_data: Dữ liệu file Excel
            
        Returns:
            DocumentContent: Nội dung đã trích xuất
        """
        try:
            # Validate file
            if not self.validate_file(file_data, '.xlsx'):
                raise ValueError("Invalid Excel file")
            
            # Xử lý theo loại file
            if file_path.lower().endswith('.xlsx'):
                return await self._process_xlsx(file_path, file_data)
            else:
                return await self._process_xls(file_path, file_data)
                
        except Exception as e:
            logger.error(f"Error extracting content from Excel: {str(e)}")
            raise
    
    async def _process_xlsx(self, file_path: str, file_data: bytes) -> DocumentContent:
        """
        Xử lý file XLSX
        
        Args:
            file_path: Đường dẫn file
            file_data: Dữ liệu file
            
        Returns:
            DocumentContent: Nội dung đã trích xuất
        """
        try:
            # Đọc Excel với pandas để lấy data
            excel_data = pd.read_excel(BytesIO(file_data), sheet_name=None, engine='openpyxl')
            
            # Đọc với openpyxl để lấy ảnh và metadata
            workbook = openpyxl.load_workbook(BytesIO(file_data))
            
            # Trích xuất text content từ tất cả sheets
            text_content = await self._extract_text_from_sheets(excel_data)
            
            # Trích xuất ảnh từ workbook (nếu process_images=True)
            images = []
            if self.process_images:
                images = await self._extract_images_from_xlsx(workbook)
            
            # Tạo metadata
            metadata = self.extract_metadata(file_path, file_data)
            metadata.update(await self._extract_xlsx_metadata(workbook, excel_data))
            
            # Convert to new page-aware structure
            return self._convert_legacy_to_page_structure(text_content, images, metadata)
            
        except Exception as e:
            logger.error(f"Error processing XLSX: {str(e)}")
            raise
    
    async def _process_xls(self, file_path: str, file_data: bytes) -> DocumentContent:
        """
        Xử lý file XLS cũ
        
        Args:
            file_path: Đường dẫn file
            file_data: Dữ liệu file
            
        Returns:
            DocumentContent: Nội dung đã trích xuất
        """
        try:
            # Đọc XLS với pandas
            excel_data = pd.read_excel(BytesIO(file_data), sheet_name=None, engine='xlrd')
            
            # Trích xuất text content
            text_content = await self._extract_text_from_sheets(excel_data)
            
            # Tạo metadata
            metadata = self.extract_metadata(file_path, file_data)
            metadata.update({
                "sheet_count": len(excel_data),
                "sheet_names": list(excel_data.keys()),
                "note": "XLS file - limited image extraction support"
            })
            
            # Convert to new page-aware structure
            return self._convert_legacy_to_page_structure(text_content, [], metadata)
            
        except Exception as e:
            logger.error(f"Error processing XLS: {str(e)}")
            raise
    
    async def _extract_text_from_sheets(self, excel_data: Dict[str, pd.DataFrame]) -> str:
        """
        Trích xuất text từ tất cả sheets
        
        Args:
            excel_data: Dictionary chứa data của các sheet
            
        Returns:
            str: Text content
        """
        text_content = ""
        
        try:
            for sheet_name, df in excel_data.items():
                text_content += f"\n=== SHEET: {sheet_name} ===\n\n"
                
                # Kiểm tra dataframe có dữ liệu không
                if df.empty:
                    text_content += "[Empty sheet]\n\n"
                    continue
                
                # Làm sạch dataframe
                df_clean = await self._clean_dataframe(df)
                
                # Convert thành text format
                sheet_text = await self._dataframe_to_text(df_clean, sheet_name)
                text_content += sheet_text + "\n\n"
            
            return self.clean_text(text_content)
            
        except Exception as e:
            logger.error(f"Error extracting text from sheets: {str(e)}")
            return ""
    
    async def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Làm sạch dataframe
        
        Args:
            df: DataFrame cần làm sạch
            
        Returns:
            pd.DataFrame: DataFrame đã làm sạch
        """
        try:
            # Loại bỏ rows và columns hoàn toàn trống
            df_clean = df.dropna(how='all', axis=0).dropna(how='all', axis=1)
            
            # Fill NaN values với empty string
            df_clean = df_clean.fillna('')
            
            # Convert tất cả về string và làm sạch
            for col in df_clean.columns:
                df_clean[col] = df_clean[col].astype(str).str.strip()
            
            return df_clean
            
        except Exception as e:
            logger.error(f"Error cleaning dataframe: {str(e)}")
            return df
    
    async def _dataframe_to_text(self, df: pd.DataFrame, sheet_name: str) -> str:
        """
        Convert DataFrame thành text format dễ đọc
        
        Args:
            df: DataFrame
            sheet_name: Tên sheet
            
        Returns:
            str: Text representation
        """
        try:
            text_parts = []
            
            # Thêm header nếu có
            if not df.columns.empty:
                headers = [str(col) for col in df.columns if str(col) != 'Unnamed: 0']
                if headers:
                    text_parts.append("Headers: " + " | ".join(headers))
            
            # Thêm data rows
            for index, row in df.iterrows():
                row_data = []
                for col in df.columns:
                    cell_value = str(row[col]).strip()
                    if cell_value and cell_value != 'nan':
                        row_data.append(cell_value)
                
                if row_data:
                    text_parts.append(" | ".join(row_data))
            
            # Thêm thống kê cơ bản
            text_parts.append(f"\n[Sheet Statistics]")
            text_parts.append(f"Rows: {len(df)}")
            text_parts.append(f"Columns: {len(df.columns)}")
            
            return "\n".join(text_parts)
            
        except Exception as e:
            logger.error(f"Error converting dataframe to text: {str(e)}")
            return f"[Error processing sheet {sheet_name}]"
    
    async def _extract_images_from_xlsx(self, workbook) -> List[Dict[str, Any]]:
        """
        Trích xuất ảnh từ file XLSX
        
        Args:
            workbook: Openpyxl workbook object
            
        Returns:
            List[Dict]: Danh sách ảnh
        """
        images = []
        image_counter = 0
        
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Tìm ảnh trong worksheet
                if hasattr(worksheet, '_images'):
                    for img in worksheet._images:
                        try:
                            # Lấy image data
                            img_data = img.ref.getvalue()
                            
                            # Validate ảnh
                            is_valid, img_metadata = DocumentUtils.validate_image_data(img_data)
                            
                            if is_valid:
                                # Tối ưu hóa ảnh
                                optimized_img_data = DocumentUtils.optimize_image(img_data)
                                
                                # Lấy vị trí ảnh nếu có
                                position = 0
                                if hasattr(img, 'anchor'):
                                    # Estimate position từ cell anchor
                                    anchor = img.anchor
                                    if hasattr(anchor, '_from'):
                                        position = anchor._from.row * 100 + anchor._from.col
                                
                                images.append({
                                    "image_data": optimized_img_data,
                                    "position": position,
                                    "page": 1,  # Excel không có page concept
                                    "image_index": image_counter,
                                    "metadata": {
                                        "sheet_name": sheet_name,
                                        "format": "Excel embedded image",
                                        **img_metadata
                                    }
                                })
                                
                                image_counter += 1
                                
                        except Exception as e:
                            logger.warning(f"Error extracting image from sheet {sheet_name}: {str(e)}")
                            continue
                            
        except Exception as e:
            logger.error(f"Error extracting images from XLSX: {str(e)}")
        
        return images
    
    async def _extract_xlsx_metadata(self, workbook, excel_data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """
        Trích xuất metadata từ XLSX
        
        Args:
            workbook: Openpyxl workbook
            excel_data: Data từ pandas
            
        Returns:
            Dict[str, Any]: Metadata
        """
        try:
            properties = workbook.properties
            
            metadata = {
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "active_sheet": workbook.active.title if workbook.active else "",
                "title": properties.title or "",
                "creator": properties.creator or "",
                "description": properties.description or "",
                "subject": properties.subject or "",
                "keywords": properties.keywords or "",
                "category": properties.category or "",
                "created": properties.created.isoformat() if properties.created else "",
                "modified": properties.modified.isoformat() if properties.modified else "",
                "last_modified_by": properties.lastModifiedBy or ""
            }
            
            # Thêm thống kê về data
            total_rows = 0
            total_columns = 0
            
            for sheet_name, df in excel_data.items():
                total_rows += len(df)
                total_columns += len(df.columns)
            
            metadata.update({
                "total_data_rows": total_rows,
                "total_data_columns": total_columns,
                "has_formulas": await self._check_has_formulas(workbook),
                "has_charts": await self._check_has_charts(workbook)
            })
            
            return metadata
            
        except Exception as e:
            logger.error(f"Error extracting XLSX metadata: {str(e)}")
            return {}
    
    async def _check_has_formulas(self, workbook) -> bool:
        """Kiểm tra có formula trong workbook không"""
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula
                            return True
            return False
        except:
            return False
    
    async def _check_has_charts(self, workbook) -> bool:
        """Kiểm tra có chart trong workbook không"""
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                if hasattr(worksheet, '_charts') and worksheet._charts:
                    return True
            return False
        except:
            return False
